from dataclasses import dataclass
from openpyxl import load_workbook
from fuzzywuzzy import fuzz

# todo get this database
workbook = load_workbook(filename="weight database.xlsx")

sheet = workbook.active
foodlist = []


@dataclass()
class Foods:
    name: str
    weight: float


for row in sheet.iter_rows(values_only=True):
    if row[0] is not None:
        food = Foods(name=row[0], weight=row[1])
    else:
        break
    foodlist.append(food)


def get_weight(needs_weight):
    score = 0
    bestmatch = ""

    for data in foodlist:
        token_set_ratio = fuzz.token_set_ratio(needs_weight, data.name.lower())
        if token_set_ratio > 10 and token_set_ratio > score:
            bestmatch = data.weight
            print(data)
            score = token_set_ratio
    return bestmatch
